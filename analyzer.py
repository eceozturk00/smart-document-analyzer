"""
Smart Document Analyzer
- Reads .docx or .pdf
- Splits text into sections using headings (best-effort)
- Extracts keywords (basic NLP)
- Exports results to Excel and JSON

Usage:
  python analyzer.py --input input.docx --excel output.xlsx --json output.json
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from docx import Document
import pdfplumber

import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize


@dataclass
class Chunk:
    section: str
    kind: str        # HEADING / TEXT
    text: str


def ensure_nltk() -> None:
    """Downloads NLTK resources if missing."""
    try:
        _ = stopwords.words("english")
    except LookupError:
        nltk.download("stopwords")
    try:
        _ = word_tokenize("test")
    except LookupError:
        nltk.download("punkt")


def clean_text(s: str) -> str:
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def infer_heading_level(style_name: str) -> Optional[int]:
    if not style_name:
        return None
    s = style_name.strip().lower()

    # English
    if "heading" in s:
        for token in s.split():
            if token.isdigit():
                return int(token)

    # Turkish variants
    if "başlık" in s or "baslik" in s:
        s = s.replace("başlık", "baslik")
        for token in s.split():
            if token.isdigit():
                return int(token)

    return None


def read_docx(path: str) -> List[Chunk]:
    doc = Document(path)
    chunks: List[Chunk] = []
    current_section = "General"

    for p in doc.paragraphs:
        text = clean_text(p.text or "")
        if not text:
            continue

        style_name = getattr(p.style, "name", "") if p.style else ""
        level = infer_heading_level(style_name)

        if level is not None:
            current_section = text
            chunks.append(Chunk(section=current_section, kind="HEADING", text=text))
        else:
            chunks.append(Chunk(section=current_section, kind="TEXT", text=text))

    return chunks


def read_pdf(path: str) -> List[Chunk]:
    chunks: List[Chunk] = []
    current_section = "General"

    with pdfplumber.open(path) as pdf:
        for page_i, page in enumerate(pdf.pages, start=1):
            raw = page.extract_text() or ""
            lines = [clean_text(l) for l in raw.splitlines() if clean_text(l)]

            for line in lines:
                # Heuristic: treat lines that look like headings as section titles
                is_heading = (
                    (len(line) <= 80 and line.isupper()) or
                    bool(re.match(r"^\d+(\.\d+)*\s+\S+", line))
                )
                if is_heading:
                    current_section = line
                    chunks.append(Chunk(section=current_section, kind="HEADING", text=line))
                else:
                    chunks.append(Chunk(section=current_section, kind="TEXT", text=line))

    return chunks


def extract_keywords(chunks: List[Chunk], top_k: int = 15) -> List[Tuple[str, int]]:
    ensure_nltk()

    text = " ".join(c.text for c in chunks if c.kind == "TEXT")
    text = re.sub(r"[^A-Za-z0-9ğüşöçıİĞÜŞÖÇ\s]", " ", text.lower())

    tokens = [t for t in word_tokenize(text) if len(t) >= 3]

    # stopwords: English + some Turkish basics
    sw = set(stopwords.words("english"))
    sw |= {"ve", "ile", "ama", "fakat", "bu", "şu", "bir", "için", "daha", "çok", "gibi", "olan", "olarak"}

    filtered = [t for t in tokens if t not in sw and not t.isdigit()]

    freq: Dict[str, int] = {}
    for w in filtered:
        freq[w] = freq.get(w, 0) + 1

    return sorted(freq.items(), key=lambda x: x[1], reverse=True)[:top_k]


def export_excel(chunks: List[Chunk], excel_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis"

    headers = ["Section", "Type", "Text"]
    ws.append(headers)

    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ch in chunks:
        ws.append([ch.section, ch.kind, ch.text])
        r = ws.max_row

        if ch.kind == "HEADING":
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=2).font = Font(bold=True)
            ws.cell(row=r, column=3).font = Font(bold=True, size=13)
        else:
            ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.column_dimensions[get_column_letter(1)].width = 32
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 90

    wb.save(excel_path)


def export_json(chunks: List[Chunk], keywords: List[Tuple[str, int]], json_path: str) -> None:
    out = {
        "stats": {
            "total_chunks": len(chunks),
            "total_text_chunks": sum(1 for c in chunks if c.kind == "TEXT"),
            "total_heading_chunks": sum(1 for c in chunks if c.kind == "HEADING"),
        },
        "keywords": [{"term": k, "count": v} for k, v in keywords],
        "chunks": [{"section": c.section, "type": c.kind, "text": c.text} for c in chunks],
    }

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)


def main() -> None:
    parser = argparse.ArgumentParser(description="Smart Document Analyzer (Word/PDF -> Excel/JSON)")
    parser.add_argument("--input", "-i", required=True, help="Input .docx or .pdf file")
    parser.add_argument("--excel", default="output.xlsx", help="Excel output path")
    parser.add_argument("--json", default="output.json", help="JSON output path")
    parser.add_argument("--topk", type=int, default=15, help="Number of keywords")
    args = parser.parse_args()

    if args.input.lower().endswith(".docx"):
        chunks = read_docx(args.input)
    elif args.input.lower().endswith(".pdf"):
        chunks = read_pdf(args.input)
    else:
        raise SystemExit(" Unsupported file type. Please provide a .docx or .pdf file.")

    if not chunks:
        raise SystemExit(" No readable content found in the document.")

    keywords = extract_keywords(chunks, top_k=args.topk)

    export_excel(chunks, args.excel)
    export_json(chunks, keywords, args.json)

    print(f" Done!\n- Excel: {args.excel}\n- JSON: {args.json}\n- Keywords: {keywords[:5]}")


if __name__ == "__main__":
    main()
